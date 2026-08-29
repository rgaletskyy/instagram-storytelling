"""Product catalogue lookup.

SKUs are named inline in the topic brief rather than passed separately, so they
are pulled out of the text and then validated against the sheet.
"""

from __future__ import annotations

import html
import mimetypes
import re
from pathlib import Path

import httpx
from openpyxl import load_workbook

from .config import IMAGE_SUFFIXES, INPUT_DIR, PRODUCTS_XLSX
from .models import Product

# Catalogue SKUs look like BO-FIU150 / ND-PAWS-01.
SKU_PATTERN = re.compile(r"\b[A-Z]{2,4}-[A-Z0-9]+(?:-[A-Z0-9]+)*\b")

_TAG_RE = re.compile(r"<[^>]+>")
_WS_RE = re.compile(r"\s+")

COLUMNS = {
    "sku": "Артикул",
    "name": "Название (UA)",
    "price": "Цена",
    "image_url": "Фото",
    "product_url": "Ссылка",
    "description": "Описание товара (UA)",
}


def strip_html(raw: object) -> str:
    """Turn the catalogue's escaped HTML description into plain text."""
    if raw is None:
        return ""
    text = html.unescape(str(raw))
    text = _TAG_RE.sub(" ", text)
    text = html.unescape(text)
    return _WS_RE.sub(" ", text).strip()


def extract_skus(brief: str) -> list[str]:
    """Find candidate SKUs in the brief, in order, without duplicates.

    Candidates are validated against the sheet by get_products, so an ordinary
    uppercase word that happens to match cannot produce a phantom product.
    """
    seen: list[str] = []
    for match in SKU_PATTERN.findall(brief or ""):
        if match not in seen:
            seen.append(match)
    return seen


def _load_rows(xlsx: Path) -> tuple[dict[str, int], list[tuple]]:
    workbook = load_workbook(xlsx, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return {}, []
    header = {str(name).strip(): i for i, name in enumerate(rows[0]) if name}
    return header, rows[1:]


def get_products(
    skus: list[str], xlsx: Path | None = None
) -> tuple[list[Product], list[str]]:
    """Look SKUs up in the catalogue.

    Returns the products found and the SKUs that had no row -- a miss is reported,
    never raised.
    """
    if not skus:
        return [], []

    header, rows = _load_rows(xlsx or PRODUCTS_XLSX)
    if not header:
        return [], list(skus)

    def cell(row: tuple, field: str) -> object:
        idx = header.get(COLUMNS[field])
        return row[idx] if idx is not None and idx < len(row) else None

    by_sku: dict[str, tuple] = {}
    for row in rows:
        raw = cell(row, "sku")
        if raw:
            by_sku[str(raw).strip().casefold()] = row

    found: list[Product] = []
    missing: list[str] = []
    for sku in skus:
        row = by_sku.get(sku.strip().casefold())
        if row is None:
            missing.append(sku)
            continue
        found.append(
            Product(
                sku=str(cell(row, "sku") or sku).strip(),
                name=str(cell(row, "name") or "").strip(),
                price=str(cell(row, "price") or "").strip(),
                image_url=str(cell(row, "image_url") or "").strip(),
                product_url=str(cell(row, "product_url") or "").strip(),
                description=strip_html(cell(row, "description")),
            )
        )
    return found, missing


def download_product_image(product: Product, out_dir: Path) -> Path | None:
    """Fetch the packshot named in the catalogue.

    Returns None when the row has no image URL or the fetch fails; callers fall
    back to a local photo. Roughly a sixth of the catalogue has an empty photo
    column, so a missing URL is ordinary, not exceptional.
    """
    if not product.image_url.startswith(("http://", "https://")):
        return None

    out_dir.mkdir(parents=True, exist_ok=True)
    try:
        response = httpx.get(product.image_url, timeout=30, follow_redirects=True)
        response.raise_for_status()
    except httpx.HTTPError:
        return None

    media_type = response.headers.get("content-type", "").split(";")[0].strip()
    if not media_type.startswith("image/"):
        return None

    suffix = mimetypes.guess_extension(media_type) or ".jpg"
    path = out_dir / f"{product.sku}{suffix}"
    path.write_bytes(response.content)
    return path


def local_product_photos(input_dir: Path | None = None, limit: int = 2) -> list[Path]:
    """Photos supplied in the input folder, as a fallback packshot source."""
    directory = input_dir or INPUT_DIR
    if not directory.exists():
        return []
    return sorted(
        p
        for p in directory.iterdir()
        if p.is_file() and p.suffix.lower() in IMAGE_SUFFIXES
    )[:limit]
