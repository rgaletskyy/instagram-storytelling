"""Product catalogue lookup.

SKUs are named inline in the topic brief rather than passed separately, so they
are pulled out of the text and then validated against the sheet.
"""

from __future__ import annotations

import html
import re
from pathlib import Path

from openpyxl import load_workbook

from .config import PRODUCTS_XLSX
from .models import Product

# Catalogue SKUs look like BO-FIU150 / ND-PAWS-01.
SKU_PATTERN = re.compile(r"\b[A-Z]{2,4}-[A-Z0-9]+(?:-[A-Z0-9]+)*\b")

_TAG_RE = re.compile(r"<[^>]+>")
_WS_RE = re.compile(r"\s+")

COLUMNS = {
    "sku": "Артикул",
    "name": "Название (UA)",
    "price": "Цена",
    "image_url": "Фото",
    "product_url": "Ссылка",
    "description": "Описание товара (UA)",
}


def strip_html(raw: object) -> str:
    """Turn the catalogue's escaped HTML description into plain text."""
    if raw is None:
        return ""
    text = html.unescape(str(raw))
    text = _TAG_RE.sub(" ", text)
    text = html.unescape(text)
    return _WS_RE.sub(" ", text).strip()


def extract_skus(brief: str) -> list[str]:
    """Find candidate SKUs in the brief, in order, without duplicates.

    Candidates are validated against the sheet by get_products, so an ordinary
    uppercase word that happens to match cannot produce a phantom product.
    """
    seen: list[str] = []
    for match in SKU_PATTERN.findall(brief or ""):
        if match not in seen:
            seen.append(match)
    return seen


def _load_rows(xlsx: Path) -> tuple[dict[str, int], list[tuple]]:
    workbook = load_workbook(xlsx, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return {}, []
    header = {str(name).strip(): i for i, name in enumerate(rows[0]) if name}
    return header, rows[1:]


def get_products(
    skus: list[str], xlsx: Path | None = None
) -> tuple[list[Product], list[str]]:
    """Look SKUs up in the catalogue.

    Returns the products found and the SKUs that had no row -- a miss is reported,
    never raised.
    """
    if not skus:
        return [], []

    header, rows = _load_rows(xlsx or PRODUCTS_XLSX)
    if not header:
        return [], list(skus)

    def cell(row: tuple, field: str) -> object:
        idx = header.get(COLUMNS[field])
        return row[idx] if idx is not None and idx < len(row) else None

    by_sku: dict[str, tuple] = {}
    for row in rows:
        raw = cell(row, "sku")
        if raw:
            by_sku[str(raw).strip().casefold()] = row

    found: list[Product] = []
    missing: list[str] = []
    for sku in skus:
        row = by_sku.get(sku.strip().casefold())
        if row is None:
            missing.append(sku)
            continue
        found.append(
            Product(
                sku=str(cell(row, "sku") or sku).strip(),
                name=str(cell(row, "name") or "").strip(),
                price=str(cell(row, "price") or "").strip(),
                image_url=str(cell(row, "image_url") or "").strip(),
                product_url=str(cell(row, "product_url") or "").strip(),
                description=strip_html(cell(row, "description")),
            )
        )
    return found, missing
